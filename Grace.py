import discord
import asyncio
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import random
import openpyxl
import datetime

BETA=True

client = discord.Client()
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

url='https://docs.google.com/spreadsheets/d/1gfSsgM_0BVqnZ02ZwRsDniU-qkRF0Wo-B7rJhYoYXqc/edit?usp=drive_web&ouid=108946956826520256706'

current_time=lambda:datetime.datetime.utcnow()+datetime.timedelta(hours=9)

@client.event
async def on_ready():
    global grace
    await client.wait_until_ready()
    grace=client.get_guild(359714850865414144)
    print("login: Grace Main")
    print(client.user.name)
    print(client.user.id)
    print("---------------")
    await client.change_presence(activity=discord.Game(name='>>', type=1))

async def get_spreadsheet(ws_name):
    creds=ServiceAccountCredentials.from_json_keyfile_name("Grace-defe42f05ec3.json", scope)
    auth=gspread.authorize(creds)

    if creds.access_token_expired:
        auth.login()
    
    try:
        worksheet=auth.open_by_url(url).worksheet(ws_name)
    except gspread.exceptions.APIError:
        return
    return worksheet

def has_role(member, role):
    return role in map(lambda x:x.name, member.roles)

async def get_member_by_battletag(battletag):
    global grace
    grace=client.get_guild(359714850865414144)

    for member in grace.members:
        try:
            if member.nick.startswith(battletag+'/'):
                return member
        except:
            continue

@client.event
async def on_message(message):
    author = message.author
    content = message.content
    channel = message.channel

    if not ((BETA and channel.id == 486550288686120961) or (not BETA and channel.id == 419397742025113612)): return

    print('{} / {}: {}'.format(channel, author, content))
    
    if message.content.startswith(">>"):
        author = message.content
        author = author.split(">>")
        author = author[1]

        if author=='':
            return
        
        spreadsheet=await get_spreadsheet('responses')
        roles=spreadsheet.col_values(6)
        battletags=spreadsheet.col_values(2)
        
        if author=="운영진":
            spreadsheet=await get_spreadsheet('staff')
            data=spreadsheet.get_all_values()
            log = '\n\n'.join(map(lambda x:'\n'.join([t for t in x if t!='']), data))
            embed = discord.Embed(title=":fire: 운영진 목록\n", description=log, color=0x5c0bb7)
            await channel.send(embed=embed)
            return

        nickname = spreadsheet.col_values(3)
        
        try:
            index = nickname.index(author) + 1
            print(index)
        except gspread.exceptions.CellNotFound:
            return
        except gspread.exceptions.APIError:
            return
        
        mention=spreadsheet.cell(index, 1).value
        battletag = spreadsheet.cell(index, 2).value
        link = spreadsheet.cell(index, 4).value
        description = spreadsheet.cell(index, 5).value
        imagelink = spreadsheet.cell(index, 6).value
        thumbnaillink = spreadsheet.cell(index, 7).value
        arena = spreadsheet.cell(index, 8).value
        league_first = spreadsheet.cell(index, 9).value
        league_second = spreadsheet.cell(index, 10).value
        friends = spreadsheet.cell(index, 11).value
        print(index, battletag, link, description, imagelink, thumbnaillink, arena, league_first, league_second, friends)

        member=await get_member_by_battletag(battletag)
        if member==None:
            return
        elif has_role(member, '클랜 마스터'):
            role='클랜 마스터'
        elif has_role(member, '운영진'):
            role='운영진'
        elif has_role(member, '클랜원'):
            role='클랜원'
        elif has_role(member, '신입 클랜원'):
            role='신입 클랜원'
        else:
            return

        print(battletag)
        print(role)
        if role == "클랜 마스터":
            roleimage = ":pen_ballpoint:"
        elif role=="운영진":
            roleimage = ":construction_worker:"
        elif role == "클랜원":
            roleimage = ":boy:"
        elif role == "신입 클랜원":
            roleimage = ":baby:"

        banned=["X", '', 'x']
        if link in banned:
            embed = discord.Embed(title="한줄소개", description=description, color=0x5c0bb7)
        else:
            embed = discord.Embed(title="바로가기", url=link, description=description, color=0x5c0bb7)

        embed.set_author(name=battletag)
        embed.add_field(name="멘션", value=mention, inline=True)
        embed.add_field(name="직책", value=roleimage + role, inline=True)
        if arena not in banned:
            embed.add_field(name="Grace Arena", value=":trophy: 제" + arena + "회 우승", inline=False)
        if league_first not in banned:
            embed.add_field(name="Grace League", value=":first_place: 제" + league_first + "회 우승", inline=False)
        if league_second not in banned:
            embed.add_field(name="Grace League", value=":second_place:제" + league_second + "회 준우승", inline=False)
        if friends not in banned:
            embed.add_field(name="우친바", value=friends, inline=False)
        if imagelink not in banned:
            embed.set_image(url=imagelink)
        if thumbnaillink not in banned:
            embed.set_thumbnail(url=thumbnaillink)

        await channel.send(embed=embed)

@client.event
async def on_message_delete(message):
    if BETA: return

    create = str(message.created_at).split('.')[0]
    if message.edited_at:
        create+='(최종수정 {})'.format(str(message.edited_at+datetime.timedelta(hours=9)).split('.')[0])
    author = message.author
    content = message.clean_content
    channel = message.channel
    delchannel = message.guild.get_channel(527859699702562828)
    await delchannel.send('{} - {} / {}: {}'.format(create, channel, author, content))

@client.event
async def on_member_join(member):
    if BETA: return

    fmt = '<@&617396702005035024>\n{0.mention}님이 {1.name}에 입장하였습니다.'
    channel = member.guild.get_channel(516122942896078868)
    role = member.guild.get_role(510731224654938112)
    await member.add_roles(role)
    await channel.send(fmt.format(member, member.guild))

@client.event
async def on_member_remove(member):
    if BETA: return

    channel = member.guild.get_channel(516122942896078868)
    fmt = '{0.mention}\n{0.nick}님이 서버에서 나가셨습니다.'
    await channel.send(fmt.format(member, member.guild))

async def periodic_sweep():
    if BETA: pass#return

    global grace
    await client.wait_until_ready()
    cur=current_time()
    next_notify=datetime.datetime(cur.year, cur.month, cur.day, 1, 51, 0)+datetime.timedelta(days=1)
    while True:
        await asyncio.sleep((next_notify-current_time()).seconds)
        next_notify+=datetime.timedelta(days=1)
        print('next sweep:', next_notify)
        
        creds=ServiceAccountCredentials.from_json_keyfile_name("Grace-defe42f05ec3.json", scope)
        auth=gspread.authorize(creds)
        
        if creds.access_token_expired:
            auth.login()

        sheet=auth.open_by_url("https://docs.google.com/spreadsheets/d/1gfSsgM_0BVqnZ02ZwRsDniU-qkRF0Wo-B7rJhYoYXqc/edit#gid=174260089")
        try:
            worksheet=sheet.worksheet('Copy of responses')
        except gspread.exceptions.APIError:
            print('spreadsheet error; trying tomorrow.')
        
        grace=client.get_guild(359714850865414144)
        res=worksheet.get_all_values()
        nicks={*map(lambda x:x.nick.split('/')[0] if (x.nick!=None and '/' in x.nick) else '', grace.members)}

        print("Command sweep")
        for i in range(1,len(res)):
            print(res[i][1], (res[i][1] not in nicks) and res[i][2]!="")
            if (res[i][1] not in nicks) and res[i][2]!="":
                worksheet.update_cell(i+1,3,"")

        print("Record sweep")
        to_be_deleted=[]
        for i in range(1,len(res)):
            print(res[i][1], (res[i][1] not in nicks), (res[i][7]+res[i][8]+res[i][9]+res[i][10]).strip()=="")
            if (res[i][1] not in nicks) and (res[i][7]+res[i][8]+res[i][9]+res[i][10]).strip()=="":
                to_be_deleted.append(i)

        print(to_be_deleted)
        for i in reversed(to_be_deleted):
            worksheet.delete_row(i)

        print('sweep finished')

access_token = os.environ["BOT_TOKEN"]
client.loop.create_task(periodic_sweep())
client.run(access_token)
