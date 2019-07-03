import discord
from discord.ext.commands import Bot
import asyncio
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import random
import openpyxl
import datetime
from time import sleep

daily=2000

gamble_channels=[486550288686120961]
gamble_notify=486550288686120961
ws_name='Beta'

content=lambda ctx:ctx.message.content
author=lambda ctx:ctx.message.author
channel=lambda ctx:ctx.message.channel.id
current_time=lambda:datetime.datetime.utcnow()+datetime.timedelta(hours=9)

client=Bot(command_prefix=('>',))
scope=['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
grace=None

@client.event
async def on_ready():
    print("login: Grace Gamble")
    print(client.user.name)
    print(client.user.id)
    print("---------------")
    await client.change_presence(activity=discord.Game(name='>>', type=1))

async def get_spreadsheet():
    creds=ServiceAccountCredentials.from_json_keyfile_name("Grace-defe42f05ec3.json", scope)
    auth=gspread.authorize(creds)

    if creds.access_token_expired:
        auth.login()

    sheet=auth.open_by_url("https://docs.google.com/spreadsheets/d/1y1XnmgggAxVVJ3jJrVBocGTjpBR7b8_L9sf47GKBNok/edit#gid=0")
    try:
        worksheet=sheet.worksheet(ws_name)
    except gspread.exceptions.APIError:
        for gamble_channel in gamble_channels:
            await client.get_channel(gamble_channel).send("API 호출 횟수에 제한이 걸렸습니다. 제발 진정하시고 잠시후 다시 시도해주세요.")
        return
    return worksheet

async def get_row(ws,user=None,mention=None):
    if user!=None:
        mention=user.mention
    if not (mention.startswith('<@') and mention.endswith('>')):
        return -1
    if mention[2]!='!':
        mention=mention[:2]+'!'+mention[2:]
    try: 
        return ws.find(mention).row
    except gspread.exceptions.CellNotFound:
        ws.append_row([mention,'0'])
        return ws.find(mention).row
    except gspread.exceptions.APIError:
        for gamble_channel in gamble_channels:
            await client.get_channel(gamble_channel).send("API 호출 횟수에 제한이 걸렸습니다. 제발 진정하시고 잠시후 다시 시도해주세요.")
        return -1

async def get_money(ws,user=None,mention=None):
    if user!=None:
        row=await get_row(ws,user)
    else:
        row=await get_row(ws,mention=mention)
    if row==-1:
        return 0
    return int(ws.cell(row,2).value)

async def redeemable(ws, user=None, mention=None):
    if user!=None:
        row=await get_row(ws,user)
    else:
        row=await get_row(ws,mention=mention)
    if row==-1:
        return False
    ct=ws.cell(row,3).value
    if ct:
        time=eval(ct)
        return current_time()-time>=datetime.timedelta(days=1, minutes=-5)
    else:
        return True

async def update_money(ws, money, user=None, mention=None, checkin=False):
    if user!=None:
        row=await get_row(ws,user)
    else:
        row=await get_row(ws,mention=mention)
    if row==-1:
        return False
    ws.update_cell(row, 2, str(money))
    if checkin:
        ws.update_cell(row, 3, repr(current_time()))
    return 1

def change_maintenance_state(ws):
    if ws.cell(1,1).value=='under maintenance':
        ws.update_cell(1,1,'userid')
        return False
    else:
        ws.update_cell(1,1,'under maintenance')
        return True

def check_maintenance_state(ws):
    return ws.cell(1,1).value=='under maintenance'

@client.command()
async def 공사(message):
    if message.channel.id!=gamble_notify: return
    commander=author(message)
    ws=await get_spreadsheet()
    if '운영진' in map(lambda x:x.name, commander.roles):
        res=change_maintenance_state(ws)
        if res:
            for gamble_channel in gamble_channels:
                await client.get_channel(gamble_channel).send('그레이스 봇 보수공사 중입니다. 제발 도박을 멈춰주세요.')
        else:
            for gamble_channel in gamble_channels:
                await client.get_channel(gamble_channel).send('보수공사가 종료되었습니다.')

@client.command()
async def 출석(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    user=author(message)
    if await redeemable(ws,user):
        money=await get_money(ws,user)
        if await update_money(ws,money+daily, user, checkin=True):
            await message.channel.send("{}\n출석체크 완료!\n현재 잔고:{}G".format(user.mention, money+daily))
            return
    await message.channel.send("{} 출석체크는 24시간에 한번만 가능합니다.".format(user.mention))

@client.command()
async def 확인(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    user=author(message)
    money=await get_money(ws,user)
    await message.channel.send("{}\n잔고:{}G".format(user.mention, money))

@client.command()
async def 송금(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    sender=author(message)
    money=await get_money(ws,sender)
    msg=content(message)
    com, rcv, send, *rest=msg.split()

    if not send.isnumeric():
        await message.channel.send("{} 송금 금액은 정수여야 합니다.".format(sender.mention))
        return

    if money<int(send):
        await message.channel.send("{} 송금 금액은 소지 금액을 넘어설 수 없습니다. 현재 소지 금액: {}".format(sender.mention, money))
        return

    if not ((rcv.startswith('<@!') and rcv.endswith('>') and rcv[3:-1].isnumeric() and grace.get_member(int(rcv[3:-1]))!=None) or
            (rcv.startswith('<@') and rcv.endswith('>') and rcv[2:-1].isnumeric() and grace.get_member(int(rcv[2:-1]))!=None)):
        await message.channel.send("{} 멘션이 잘못되었습니다.".format(sender.mention))
        return

    rcv_mon=await get_money(ws,mention=rcv)
    await update_money(ws,rcv_mon+int(send), mention=rcv)
    await update_money(ws,money-int(send), sender)

    await message.channel.send("송금 완료: {} -> {}\n보낸 사람 잔고: {}G\n받는 사람 잔고: {}G".format(sender.mention, rcv, money-int(send), rcv_mon+int(send)))

@client.command()
async def 동전(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    user=author(message)
    msg=content(message)
    com, choice, bet, *rest=msg.split()
    
    if choice not in ('앞', '뒤'):
        await message.channel.send("{} 앞 또는 뒤만 선택할 수 있습니다.".format(user.mention))
        return
    
    if not bet.isnumeric():
        await message.channel.send("{} 베팅 금액은 자연수여야 합니다.".format(user.mention))
        return
    
    bet=int(bet)
    if bet==0:
        await message.channel.send("{} 베팅 금액은 자연수여야 합니다.".format(user.mention))
        return
    
    money=await get_money(ws,user)
    if (bet,money)!=(1,0) and bet>money:
        await message.channel.send("{} 베팅 금액은 소지 금액을 넘어설 수 없습니다. 현재 소지 금액: {}".format(user.mention, money))
        return

    msg="{}\n예측:{}\n동전:".format(user.mention,choice)
    
    result=random.choice(['앞','뒤'])
    msg+=result+'\n'

    if result==choice:
        msg+=':white_check_mark: 성공!\n'
        money+=bet
    else:
        msg+=':x: 실패...\n'
        if money==0:
            money=1
        money-=bet

    await update_money(ws,money, user)
    msg+='현재 잔고: {}'.format(money)

    await message.channel.send(msg)

@client.command()
async def 순위(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    user=author(message)
    money=await get_money(ws,user)

    moneys=[*sorted(map(lambda x:int(x) if x.isnumeric() else -1,ws.col_values(2)), reverse=True)]
    rank=moneys.index(money)+1
    same=moneys.count(money)
    await message.channel.send("{}\n현재 {}위(동순위 {}명)".format(user.mention, rank, same))

@client.command()
async def 랭킹(message):
    if message.channel.id not in gamble_channels: return
    ws=await get_spreadsheet()
    if check_maintenance_state(ws):
        await message.channel.send("진정하시라고요.")
        return
    user=author(message)
    msg=content(message)

    data=ws.get_all_values()[1:]
    data.sort(key=lambda x:int(x[1]), reverse=True)

    ct=msg.split()
    maxrank=min(10, len(data))
    if len(ct)>1 and ct[1].isnumeric():
        maxrank=min(int(ct[1]), len(data))

    log="현재 랭킹"
    cnt=0
    par_cnt=1
    prev_money=-1
    for d in data:
        user=grace.get_member(int(d[0][3:-1]))
        if user:
            if prev_money==int(d[1]):
                par_cnt+=1
            else:
                cnt+=par_cnt
                if cnt>maxrank or int(d[1])==0:
                    break
                par_cnt=1
                prev_money=int(d[1])
            log+="\n{}. {}: {}G".format(cnt, user.nick.split('/')[0], d[1])

    await message.channel.send(log)

@client.command()
async def 도움말(message):
    if message.channel.id not in gamble_channels: return
    embed = discord.Embed(title="Grace gamble bot", description="그레이스 클랜 도박 봇입니다.", color=0xeee657)
    embed.add_field(name=">출석\n",value="2000G를 받습니다. 23시간 55분에 한 번만 사용할 수 있습니다.\n",inline=False)
    embed.add_field(name=">확인\n",value="자신의 소지 G를 확인합니다.\n",inline=False)
    embed.add_field(name=">송금 (멘션) (금액)\n",value="멘션한 사람에게 언급된 금액을 송금합니다.\n",inline=False)
    embed.add_field(name=">동전 [앞/뒤] (금액)\n",value="G를 걸고, 동전을 던집니다. 맞추면 두 배로 돌려받고, 틀리면 돌려받지 못합니다.\n0G를 소지중이라면 1G를 걸어 성공시 1G를 받을 수 있습니다.",inline=False)
    embed.add_field(name=">순위\n",value="자신의 순위와 동순위인 사람 수를 알려줍니다.\n",inline=False)
    embed.add_field(name=">랭킹 {순위}\n",value="순위까지의 랭킹을 표시합니다. {순위}값을 주지 않거나 숫자가 아니라면 10위까지 표시합니다.\n",inline=False)
    embed.add_field(name="버그바운티\n",value="유의미한 버그를 제보주신 분들께는 인게임에서 사용하실 수 있는 G를 버그 수준에 따라 드립니다.\n",inline=False)
    await message.send(embed=embed)

async def periodic_ranking():
    global grace
    await client.wait_until_ready()
    grace=client.get_guild(359714850865414144)
    cur=current_time()
    next_notify=datetime.datetime(cur.year, cur.month, cur.day, 0, 0, 0)+datetime.timedelta(days=1)
    while True:
        await asyncio.sleep((next_notify-current_time()).seconds)

        ws=await get_spreadsheet()
        data=ws.get_all_values()[1:]
        data.sort(key=lambda x:int(x[1]), reverse=True)

        maxrank=10
        c=current_time()
        log="{}년 {}월 {}일 일일 랭킹".format(c.year, c.month, c.day)
        cnt=0
        par_cnt=1
        prev_money=-1
        for d in data:
            user=grace.get_member(int(d[0][3:-1]))
            if user:
                if prev_money==int(d[1]):
                    par_cnt+=1
                else:
                    cnt+=par_cnt
                    if cnt>maxrank or int(d[1])==0:
                        break
                    par_cnt=1
                    prev_money=int(d[1])
                log+="\n{}. {}: {}G".format(cnt, user.nick.split('/')[0], d[1])

        for gamble_channel in gamble_channels:
            await client.get_channel(gamble_channel).send(log)
        next_notify+=datetime.timedelta(days=1)

access_token = os.environ["BOT_TOKEN"]

client.loop.create_task(periodic_ranking())
client.run(access_token)
